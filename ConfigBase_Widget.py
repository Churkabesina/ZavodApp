import os.path
import sqlite3
from PyQt6 import QtCore, QtWidgets, QtGui
import shutil
import openpyxl
import re


class ConfigBaseWidget(QtWidgets.QWidget):
    def __init__(self, rows_data: list[tuple], data_path, draws_path, output_path, src_path, main_widget):
        super().__init__()

        self.truba_name = {'Труба', 'труба', 'ТРУБА'}
        self.krug_name = {'Круг', 'круг', 'КРУГ'}
        self.font_rows = QtGui.QFont('Times New Roman', 14)
        self.main_widget: QtWidgets.QMainWindow = main_widget
        self.DATA_PATH = data_path
        self.DRAWS_PATH = draws_path
        self.OUTPUT_PATH = output_path
        self.SRC_PATH = src_path
        self.setObjectName("config_base")
        self.setWindowTitle("Настройка базы данных")
        self.resize(1300, 682)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        # Кнопка импорт
        self.horizontalLayoutExportImport = QtWidgets.QHBoxLayout()
        self.import_btn = QtWidgets.QPushButton(parent=self)
        self.import_btn.setMaximumSize(150, 50)
        self.import_btn.setText('Импорт')
        self.import_btn.setFont(self.font_rows)
        self.import_btn.clicked.connect(self.import_excel)
        self.horizontalLayoutExportImport.addWidget(self.import_btn)
        # Кнопка экспорт
        self.export_btn = QtWidgets.QPushButton(parent=self)
        self.export_btn.setMaximumSize(150, 50)
        self.export_btn.setText('Экспорт')
        self.export_btn.setFont(self.font_rows)
        self.export_btn.clicked.connect(self.export_excel)
        self.horizontalLayoutExportImport.addWidget(self.export_btn)
        # Добавить кнопки на лейаут
        self.verticalLayout_2.addLayout(self.horizontalLayoutExportImport)
        # Скролл ареа
        self.scrollArea = QtWidgets.QScrollArea(parent=self)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setAlignment(
            QtCore.Qt.AlignmentFlag.AlignLeading | QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 926, 581))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout(self.scrollAreaWidgetContents)
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.verticalLayout.addWidget(self.scrollArea)
        self.verticalLayout_2.addLayout(self.verticalLayout)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        # Кнопка ОК
        self.pushButton = QtWidgets.QPushButton(parent=self)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setFont(self.font_rows)
        self.horizontalLayout.addWidget(self.pushButton)
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.pushButton.setText("OK")
        self.pushButton.setMinimumSize(QtCore.QSize(20, 40))
        self.pushButton.setMaximumSize(QtCore.QSize(400, 100))
        self.pushButton.clicked.connect(self.close_widget)

        self.frame_new_row_init()
        # self.frame_categories_init()

        for index, row in enumerate(rows_data, 1):
            self.frame_init(f'frame_{index}', row)

    def frame_categories_init(self):
        self.frame = QtWidgets.QFrame(parent=self.scrollAreaWidgetContents)
        self.frame.setMinimumSize(QtCore.QSize(0, 100))
        self.frame.setMaximumSize(QtCore.QSize(16777215, 100))
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName('category_frame')
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.frame)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        self.product = QtWidgets.QLabel(parent=self.frame)
        self.product.setFont(self.font_rows)
        self.product.setObjectName("product")
        self.product.setText('Продукция')
        self.horizontalLayout_2.addWidget(self.product)

        self.diameter = QtWidgets.QLabel(parent=self.frame)
        self.diameter.setFont(self.font_rows)
        self.diameter.setObjectName("diameter")
        self.diameter.setText('Диаметр, мм')
        self.horizontalLayout_2.addWidget(self.diameter)

        self.type_metall = QtWidgets.QLabel(parent=self.frame)
        self.type_metall.setFont(self.font_rows)
        self.type_metall.setObjectName("type_metall")
        self.type_metall.setText('Тип металла')
        self.horizontalLayout_2.addWidget(self.type_metall)

        self.type_steel = QtWidgets.QLabel(parent=self.frame)
        self.type_steel.setFont(self.font_rows)
        self.type_steel.setObjectName("type_steel")
        self.type_steel.setText('Марка стали')
        self.horizontalLayout_2.addWidget(self.type_steel)

        self.lenght = QtWidgets.QLabel(parent=self.frame)
        self.lenght.setFont(self.font_rows)
        self.lenght.setObjectName("lenght")
        self.lenght.setText('Длина, мм')
        self.horizontalLayout_2.addWidget(self.lenght)

        self.weight = QtWidgets.QLabel(parent=self.frame)
        self.weight.setObjectName("weight")
        self.weight.setFont(self.font_rows)
        self.weight.setText('Вес за единицу')
        self.horizontalLayout_2.addWidget(self.weight)

        self.draw_path = QtWidgets.QLabel(parent=self.frame)
        self.draw_path.setFont(self.font_rows)
        self.draw_path.setObjectName("draw_path")
        self.draw_path.setText('Чертеж')
        self.horizontalLayout_2.addWidget(self.draw_path)

        self.verticalLayout_3.addLayout(self.horizontalLayout_2)
        self.verticalLayout_6.addWidget(self.frame)

    def frame_new_row_init(self):
        self.frame_new_row = QtWidgets.QFrame(parent=self.scrollAreaWidgetContents)
        self.frame_new_row.setGeometry(QtCore.QRect(130, 90, 921, 100))
        self.frame_new_row.setMinimumSize(QtCore.QSize(0, 100))
        self.frame_new_row.setMaximumSize(QtCore.QSize(16777215, 100))
        self.frame_new_row.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame_new_row.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame_new_row.setObjectName("frame_new_row")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.frame_new_row)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        self.product = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.product.setObjectName("product")
        self.product.setFont(self.font_rows)
        self.horizontalLayout.addWidget(self.product)

        # self.type_metall = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.type_metall = QtWidgets.QComboBox(parent=self.frame_new_row)
        self.type_metall.addItems(['Круг', 'Труба'])
        self.type_metall.setObjectName("type_metall")
        self.type_metall.setFont(self.font_rows)
        # self.type_metall.currentIndexChanged.connect(self.type_metall_changed)
        self.horizontalLayout.addWidget(self.type_metall)

        self.diameter = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.diameter.setObjectName("diameter")
        self.diameter.setFont(self.font_rows)
        # self.diameter.addItems([str(x) for x in range(30, 351, 5)])
        # self.diameter.setPlaceholderText('')
        # self.diameter.setCurrentIndex(-1)
        self.horizontalLayout.addWidget(self.diameter)

        # self.tolshina = QtWidgets.QComboBox(parent=self.frame_new_row)
        # self.tolshina.setObjectName('tolshina')
        # self.tolshina.setFont(self.font_rows)
        # self.tolshina.addItems([str(x / 10) for x in range(5, 100, 5)])
        # self.tolshina.setPlaceholderText('')
        # self.tolshina.setCurrentIndex(-1)
        # self.tolshina.setEnabled(False)
        # self.horizontalLayout.addWidget(self.tolshina)

        self.type_steel = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.type_steel.setText("")
        self.type_steel.setFont(self.font_rows)
        self.type_steel.setObjectName("type_steel")
        self.horizontalLayout.addWidget(self.type_steel)

        self.lenght = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.lenght.setObjectName("lenght")
        self.lenght.setFont(self.font_rows)
        self.horizontalLayout.addWidget(self.lenght)

        self.draw_path = QtWidgets.QLineEdit(parent=self.frame_new_row)
        self.draw_path.setReadOnly(True)
        self.draw_path.setObjectName("draw_path")
        self.draw_path.setFont(self.font_rows)
        self.horizontalLayout.addWidget(self.draw_path)

        self.choose_path = QtWidgets.QPushButton(parent=self.frame_new_row)
        self.choose_path.setMinimumSize(QtCore.QSize(50, 0))
        self.choose_path.setObjectName("choose_path")
        self.choose_path.setIcon(QtGui.QIcon(r".\Src\folder.png"))
        self.choose_path.clicked.connect(self.choose_draw)
        self.horizontalLayout.addWidget(self.choose_path)

        self.create_new_rows = QtWidgets.QPushButton(parent=self.frame_new_row)
        self.create_new_rows.setMinimumSize(QtCore.QSize(170, 0))
        self.create_new_rows.setFont(self.font_rows)
        self.create_new_rows.setObjectName("create_new_rows")
        self.horizontalLayout.addWidget(self.create_new_rows)
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.product.setPlaceholderText("Продукт")
        self.diameter.setPlaceholderText("Диаметр, мм")
        # self.tolshina.setPlaceholderText('Толщина')
        self.type_metall.setPlaceholderText("Вид металла")
        self.type_steel.setPlaceholderText("Марка стали")
        self.lenght.setPlaceholderText("Длина, мм")
        # self.weight.setPlaceholderText("Вес, кг/м")
        self.draw_path.setPlaceholderText("Чертёж")
        self.create_new_rows.setText("Добавить запись")
        self.create_new_rows.clicked.connect(self.add_product)
        self.verticalLayout_6.addWidget(self.frame_new_row)

    def frame_init(self, frame_name: str, row_data: tuple[str]):  # Инициализация нового фрейма
        self.frame = QtWidgets.QFrame(parent=self.scrollAreaWidgetContents)
        self.frame.setMinimumSize(QtCore.QSize(0, 100))
        self.frame.setMaximumSize(QtCore.QSize(16777215, 100))
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName(frame_name)
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.frame)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.product = QtWidgets.QLineEdit(parent=self.frame)

        self.product.setFont(self.font_rows)
        self.product.setReadOnly(True)
        self.product.setObjectName("product")
        self.horizontalLayout_2.addWidget(self.product)

        self.diameter = QtWidgets.QLineEdit(parent=self.frame)
        self.diameter.setFont(self.font_rows)
        self.diameter.setReadOnly(True)
        self.diameter.setObjectName("diameter")
        self.diameter.setPlaceholderText('Диаметр, мм')
        self.horizontalLayout_2.addWidget(self.diameter)

        self.type_metall = QtWidgets.QLineEdit(parent=self.frame)
        self.type_metall.setFont(self.font_rows)
        self.type_metall.setReadOnly(True)
        self.type_metall.setObjectName("type_metall")
        self.horizontalLayout_2.addWidget(self.type_metall)

        self.type_steel = QtWidgets.QLineEdit(parent=self.frame)
        self.type_steel.setFont(self.font_rows)
        self.type_steel.setReadOnly(True)
        self.type_steel.setObjectName("type_steel")
        self.type_steel.setPlaceholderText('Марка стали')
        self.horizontalLayout_2.addWidget(self.type_steel)

        self.lenght = QtWidgets.QLineEdit(parent=self.frame)
        self.lenght.setFont(self.font_rows)
        self.lenght.setReadOnly(True)
        self.lenght.setObjectName("lenght")
        self.lenght.setPlaceholderText('Длина, мм')
        self.horizontalLayout_2.addWidget(self.lenght)

        self.weight = QtWidgets.QLineEdit(parent=self.frame)
        self.weight.setReadOnly(True)
        self.weight.setObjectName("weight")
        self.weight.setFont(self.font_rows)
        self.weight.setPlaceholderText('Вес за единицу')
        self.horizontalLayout_2.addWidget(self.weight)

        self.draw_path = QtWidgets.QLineEdit(parent=self.frame)
        self.draw_path.setFont(self.font_rows)
        self.draw_path.setReadOnly(True)
        self.draw_path.setObjectName("draw_path")
        self.draw_path.setPlaceholderText('Чертёж')
        self.horizontalLayout_2.addWidget(self.draw_path)

        self.checkbox = QtWidgets.QCheckBox(parent=self.frame)
        self.checkbox.setText("")
        self.checkbox.setObjectName("checkbox")
        self.horizontalLayout_2.addWidget(self.checkbox)
        self.update_btn = QtWidgets.QPushButton(parent=self.frame)
        self.update_btn.setEnabled(False)
        self.update_btn.setObjectName("update_btn_1")
        self.horizontalLayout_2.addWidget(self.update_btn)
        self.del_btn = QtWidgets.QPushButton(parent=self.frame)
        self.del_btn.setEnabled(False)
        self.del_btn.setMinimumSize(QtCore.QSize(10, 0))
        self.del_btn.setObjectName("del_btn")
        self.horizontalLayout_2.addWidget(self.del_btn)
        self.verticalLayout_3.addLayout(self.horizontalLayout_2)
        self.verticalLayout_6.addWidget(self.frame)
        self.update_btn.setText("обновить")
        self.update_btn.setFont(self.font_rows)
        self.del_btn.setText("del")
        self.del_btn.setFont(self.font_rows)

        self.frame_data_init(row_data)
        self.frame_signals_init()

    def frame_data_init(self, row_data: tuple):  # заполнение данными все QlineEdit
        self.product.setText(row_data[0])
        self.product.setPlaceholderText(row_data[0])
        self.diameter.setText(row_data[1])
        self.type_metall.setText(row_data[2])
        self.type_steel.setText(row_data[3])
        self.lenght.setText(row_data[4])
        self.weight.setText(row_data[5])
        self.draw_path.setText(row_data[6])
        self.del_btn.setObjectName(str(row_data[-1]))
        self.update_btn.setObjectName(str(row_data[-1]))

    def frame_signals_init(self):  # привязка сигналов к кнопкам во фрейме
        self.checkbox.toggled.connect(self.toggle_checkbox)
        self.update_btn.clicked.connect(self.update_product)
        self.del_btn.clicked.connect(self.delete_product)

    def close_widget(self):
        self.deleteLater()

    def toggle_checkbox(self):  # защитный чекбокс
        check_box: QtWidgets.QCheckBox = self.sender()
        parent_frame = check_box.parent()
        if check_box.isChecked():
            for children in parent_frame.children():
                if isinstance(children, QtWidgets.QLineEdit):
                    children.setReadOnly(False)
                if isinstance(children, QtWidgets.QPushButton):
                    children.setEnabled(True)
        else:
            for children in parent_frame.children():
                if isinstance(children, QtWidgets.QLineEdit):
                    children.setReadOnly(True)
                if isinstance(children, QtWidgets.QPushButton):
                    children.setEnabled(False)

    def update_product(self):  # обновить 1 продукцию в базе
        update_btn = self.sender()
        qline_product = update_btn.parent().findChild(QtWidgets.QLineEdit, 'product')
        old_text = qline_product.placeholderText()
        id = update_btn.objectName()
        parent_frame = update_btn.parent()
        update_data = dict()
        for children in parent_frame.children():
            if isinstance(children, QtWidgets.QLineEdit):
                update_data[children.objectName()] = children.text()
            if isinstance(children, QtWidgets.QCheckBox):
                children.setChecked(False)
                children.setEnabled(False)
            else:
                children.setEnabled(False)
        update_data['id'] = id
        with sqlite3.connect(f'{self.DATA_PATH}database.db') as update_conn:
            cursor = update_conn.cursor()
            # {'product': 'Штруцель', 'diameter': '56x20'}
            cursor.execute('UPDATE Products SET product = ?, diameter = ?, type_metall = ?, type_steel = ?,'
                           ' lenght = ?, weight = ?, draw_path = ? WHERE id = ?',
                           (update_data['product'], update_data['diameter'], update_data['type_metall'],
                            update_data['type_steel'], update_data['lenght'], update_data['weight'],
                            update_data['draw_path'],
                            update_data['id']))
            update_conn.commit()
        cursor.close()
        update_conn.close()
        for frame in self.main_widget.findChild(QtWidgets.QFrame, 'frame_for_rows').children():
            if isinstance(frame, QtWidgets.QFrame):
                combobox_main_widget = frame.findChild(QtWidgets.QComboBox)
                if combobox_main_widget.currentText() == old_text:
                    combobox_main_widget.setItemText(combobox_main_widget.findText(old_text), update_data['product'])
                    combobox_main_widget.setCurrentText('')
                    combobox_main_widget.setCurrentText(old_text)
                    qline_product.setPlaceholderText(update_data['product'])
                else:
                    combobox_main_widget.setItemText(combobox_main_widget.findText(old_text), update_data['product'])
                    qline_product.setPlaceholderText(update_data['product'])
        self.main_widget.list_products_init()

    def delete_product(self):  # удалить 1 продукцию из базы
        delete_btn = self.sender()
        product = delete_btn.parent().findChild(QtWidgets.QLineEdit, 'product').text()
        product_id = delete_btn.objectName()
        parent_frame = delete_btn.parent()
        with sqlite3.connect(f'{self.DATA_PATH}database.db') as delete_conn:
            cursor = delete_conn.cursor()
            cursor.execute('DELETE FROM Products WHERE id = ?', (product_id,))
        cursor.close()
        delete_conn.close()
        for frame in self.main_widget.findChild(QtWidgets.QFrame, 'frame_for_rows').children():
            if isinstance(frame, QtWidgets.QFrame):
                combobox = frame.findChild(QtWidgets.QComboBox)
                combobox.removeItem(combobox.findText(f'{product}'))
        self.main_widget.list_products_init()

        parent_frame.deleteLater()

    def add_product(self):
        add_btn = self.sender()
        product = add_btn.parent().findChild(QtWidgets.QLineEdit, 'product')
        diameter = add_btn.parent().parent().findChild(QtWidgets.QLineEdit, 'diameter')
        # tolshina = add_btn.parent().parent().findChild(QtWidgets.QComboBox, 'tolshina')
        type_metall = add_btn.parent().parent().findChild(QtWidgets.QComboBox, 'type_metall')
        type_steel = add_btn.parent().parent().findChild(QtWidgets.QLineEdit, 'type_steel')
        lenght = add_btn.parent().parent().findChild(QtWidgets.QLineEdit, 'lenght')
        draw_path = add_btn.parent().parent().findChild(QtWidgets.QLineEdit, 'draw_path')
        if len(product.text()) != 0:
            if diameter.text() != '' and lenght.text() != '':

                if type_metall.currentText() == 'Труба':
                    len_line_value = float(lenght.text().replace(',', '.'))
                    diameter_value = float(re.split(r'[xXхХ]', diameter.text().replace(',', '.'))[0])
                    tolshina_value = float(re.split(r'[xXхХ]', diameter.text().replace(',', '.'))[1])
                    weight = str(round(6.16225 * (diameter_value ** 2 - (diameter_value - 2 * tolshina_value) ** 2)*len_line_value/(10 ** 9) * 1000, 5))
                else:
                    len_line_value = float(lenght.text().replace(',', '.'))
                    diameter_value = float(diameter.text().replace(',', '.'))
                    weight = str(round(6.169315 * (diameter_value ** 2) * len_line_value / (10 ** 6), 5))

                with sqlite3.connect(f'{self.DATA_PATH}database.db') as add_conn:
                    cursor = add_conn.cursor()
                    cursor.execute(
                        'INSERT INTO Products (product, diameter, type_metall, type_steel, lenght, weight, draw_path) VALUES (?,?,?,?,?,?,?)',
                        (product.text(), diameter.text(), type_metall.currentText(), type_steel.text(), lenght.text(),
                         weight, draw_path.text()))
                    add_conn.commit()
                cursor.close()
                add_conn.close()
                for frame in self.main_widget.findChild(QtWidgets.QFrame, 'frame_for_rows').children():
                    if isinstance(frame, QtWidgets.QFrame):
                        frame.findChild(QtWidgets.QComboBox).addItem(f'{product.text()}')
                self.main_widget.list_products_init()
                product.setText('')
                diameter.setText('')
                type_steel.setText('')
                lenght.setText('')
                draw_path.setText('')
                path = draw_path.placeholderText()
                if path != 'Чертёж':
                    shutil.copy(path, self.DRAWS_PATH)  # !!!!!!ошибка про копировании одинаковых файлов!!!!!
                draw_path.setPlaceholderText('Чертёж')
                self.create_warn_messange('Уведомление', 'Новая запись успешно добавлена')

            else:
                with sqlite3.connect(f'{self.DATA_PATH}database.db') as add_conn:
                    cursor = add_conn.cursor()
                    cursor.execute(
                        'INSERT INTO Products (product, diameter, type_metall, type_steel, lenght, weight, draw_path) VALUES (?,?,?,?,?,?,?)',
                        (product.text(), diameter.text(), type_metall.currentText(), type_steel.text(), lenght.text(),
                         '', draw_path.text()))
                    add_conn.commit()
                cursor.close()
                add_conn.close()
                for frame in self.main_widget.findChild(QtWidgets.QFrame, 'frame_for_rows').children():
                    if isinstance(frame, QtWidgets.QFrame):
                        frame.findChild(QtWidgets.QComboBox).addItem(f'{product.text()}')
                self.main_widget.list_products_init()
                product.setText('')
                diameter.setText('')
                type_steel.setText('')
                lenght.setText('')
                draw_path.setText('')
                path = draw_path.placeholderText()
                if path != 'Чертёж':
                    shutil.copy(path, self.DRAWS_PATH)  # !!!!!!ошибка про копировании одинаковых файлов!!!!!
                draw_path.setPlaceholderText('Чертёж')
                self.create_warn_messange('Уведомление', 'Новая запись успешно добавлена')
        else:
            self.create_warn_messange('Предупреждение', 'Заполните хотя бы название', '250')

    def closeEvent(self, a0):
        self.deleteLater()

    def accepting_msg_box_config_widget(self):
        self.warn_msg.deleteLater()

    def choose_draw(self):
        pdf_path = QtWidgets.QFileDialog.getOpenFileName(
            self,
            'Выберите чертеж в формате pdf',
            filter='PDF (*.pdf)')
        if pdf_path[0] == '':
            return None
        self.frame_new_row.findChild(QtWidgets.QLineEdit, 'draw_path').setPlaceholderText(f'{pdf_path[0]}')
        path_to_pdf = pdf_path[0].split('/')
        self.frame_new_row.findChild(QtWidgets.QLineEdit, 'draw_path').setText(f'{path_to_pdf[-1]}')

    def import_excel(self):
        table_path = QtWidgets.QFileDialog.getOpenFileName(self,
                                                           caption='Выберите таблицу в формате xlsx',
                                                           filter='Excel (*.xlsx)')
        if table_path[0] != '':
            wb = openpyxl.load_workbook(table_path[0], read_only=True)  # загрузить excel
            ws = wb['Исходные данные']

            if ws.max_column != 7:
                self.create_warn_messange('Предупреждение', 'Количество столбцов в таблице неверно')
                return None

            data = [[col.value for col in row] for row in ws.iter_rows()][1:]
            main_widget_frame_for_rows_childs = self.main_widget.findChild(QtWidgets.QFrame, 'frame_for_rows').children()
            with sqlite3.connect(f'{self.DATA_PATH}database.db') as conn:
                cursor = conn.cursor()
                for row in data:

                    if row[6] is not None:
                        if row[6].strip() != '':
                            if not os.path.isfile(row[6].strip()):
                                self.create_warn_messange('Предупреждение', 'В импортируемой таблице указан неверный путь к чертежу', '350')
                                return None
                            else:
                                shutil.copy(row[6].strip(), self.DRAWS_PATH)
                                row[6] = re.split(r'[\\/]', row[6])[-1]

                    if str(row[1]).strip() == '' or str(row[4]).strip() == '' or str(row[2]).strip() == '':
                        self.create_warn_messange('Предупреждение', 'В импортируемой таблице не все строки заполнены')
                        return None

                    if row[2].strip() in self.truba_name:
                        row[1] = str(row[1]).replace(',', '.')
                        row[4] = str(row[4]).replace(',', '.')
                        diameter_value = float(re.split(r'[xXхХ]', row[1])[0])
                        tolshina_value = float(re.split(r'[xXхХ]', row[1])[1])
                        len_line_value = float(row[4])
                        row[5] = str(round(6.16225*(diameter_value**2 - ((diameter_value - 2*tolshina_value)**2))*len_line_value/(10**9) * 1000, 5))
                    else:
                        row[1] = str(row[1]).replace(',', '.')
                        row[4] = str(row[4]).replace(',', '.')
                        diameter_value = float(row[1])
                        len_line_value = float(row[4])
                        row[5] = str(round(6.169315*(diameter_value**2)*len_line_value/(10**6), 5))
                    cursor.execute(
                        'INSERT INTO Products (product, diameter, type_metall, type_steel, lenght, weight, draw_path) VALUES (?,?,?,?,?,?,?)',
                        (row[0], row[1], row[2], row[3], row[4], row[5], row[6]))
                    for frame in main_widget_frame_for_rows_childs:
                        if isinstance(frame, QtWidgets.QFrame):
                            frame.findChild(QtWidgets.QComboBox).addItem(f'{row[0]}')
            cursor.close()
            conn.close()
            self.main_widget.list_products_init()
            self.create_warn_messange('Уведомление', 'База данных импортирована')

    def export_excel(self):
        table_path = QtWidgets.QFileDialog.getExistingDirectory(self,
                                                                caption='Выберите папку для экспорта таблицы'
                                                                )
        if table_path.strip(' ') != '':
            wb = openpyxl.Workbook()
            ws = wb.active
            with sqlite3.connect(f'{self.DATA_PATH}database.db') as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT product, diameter, type_metall, type_steel, lenght, weight, draw_path from Products')
                data: tuple[tuple] = cursor.fetchall()  # ((product, diameter, type_metall, type_steel, lenght, weight, draw_path),(),())
            cursor.close()
            conn.close()
            ws.append(('Наименование', 'Диаметр, мм', 'Вид металла', 'Марка стали', 'Длина, мм', 'Вес, кг/м', 'Чертеж'))
            for row in data:
                ws.append(row)
            wb.save(f'{table_path}/exported_table.xlsx')
            self.create_warn_messange('Уведомление', 'База данных экспортирована')

    def create_warn_messange(self, title: str, text: str, size: str = '250') -> None:
        self.warn_msg = QtWidgets.QMessageBox()
        self.warn_msg.setWindowTitle(f'{title}')
        self.warn_msg.setStyleSheet(f'QLabel {{ min-width: {size}px}}')
        self.warn_msg.setText(f'{text}')
        self.warn_msg.accepted.connect(self.accepting_msg_box_config_widget)
        self.warn_msg.show()

    # def type_metall_changed(self):
    #     if self.sender().currentText() == 'Труба':
    #         print('Truba')
    #         # self.tolshina.setEnabled(True)
    #     else:
    #         print('Krug')
    #         # self.tolshina.setEnabled(False)
    #         # self.tolshina.setCurrentIndex(-1)